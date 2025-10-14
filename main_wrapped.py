from parse_utils import parse_first_table, parse_second_table
from openpyxl import load_workbook
from tkinter import Tk, filedialog
import pandas as pd
from logger import logger

def process_documents():
    # Открываем диалог выбора файлов
    root = Tk()
    root.withdraw()  # скрываем основное окно

    # выбор файла для сохранения данных
    output_path = filedialog.asksaveasfilename(
        initialdir="C:/COTECNA",
        defaultextension=".xlsx",
        filetypes=[("Excel файлы", "*.xlsx")],
        title="Сохранить как"
    )

    if not output_path:
        logger.info("Файлы не выбраны.")
        return

    # Выбор нескольких .docx файлов
    file_paths = filedialog.askopenfilenames(
        initialdir="C:/COTECNA",
        title="Выберите файлы",
        filetypes=[("Word документы", "*.docx")]
    )

    if not file_paths:
        logger.info("Файлы не выбраны.")
        return

    sheet_name = "Результаты анализа"

    for doc_path in file_paths:
        # Определяем номер строки, с которой надо начинать запись
        book = load_workbook(output_path)
        sheet = book[sheet_name]
        startrow = 1
        while sheet[f"A{startrow}"].value is not None:
            startrow += 1

        # info_df = parse_first_table(doc_path)
        # results_df = parse_second_table(doc_path)
        # df_to_add = pd.concat([info_df, results_df], axis=1)

        try:
            info_df = parse_first_table(doc_path)
            results_df = parse_second_table(doc_path)

            if info_df.empty and results_df.empty:
                logger.warning(f"⚠️ Файл '{doc_path}' не содержит пригодных данных (пустые таблицы). Пропущен.")
                continue

            df_to_add = pd.concat([info_df, results_df], axis=1)

        except Exception as e:
            logger.error(f"⛔ Ошибка при разборе Word-файла '{doc_path}': {e}")
            continue  # пропускаем файл


        existing_headers = [cell.value for cell in sheet[1] if cell.value is not None]
        # Определяем "лишние" колонки, которые есть в df, но отсутствуют в Excel
        missing_columns = [col for col in df_to_add.columns if col not in existing_headers]

        if missing_columns:
            # logger.warning(
            #     f"Пропущенные колонки для '{doc_path.split('/')[-1]}' в листе '{sheet_name}': {missing_columns}"
            # )
            # Сохраняем эти колонки и их значения в отдельный лист
            new_data = df_to_add[missing_columns].copy()
            new_data.insert(0, "Источник", doc_path.split("/")[-1])  # Добавим имя файла для контекста

            if "Новые поля" not in book.sheetnames:
                book.create_sheet("Новые поля")

            new_sheet = book["Новые поля"]
            new_sheet_start = 1
            while new_sheet[f"A{new_sheet_start}"].value is not None:
                new_sheet_start += 1

            with pd.ExcelWriter(output_path, engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:
                # Таблица анализа
                new_data.to_excel(writer, sheet_name="Новые поля", index=False, startrow=new_sheet_start - 1, header=True)

        # Добавим отсутствующие колонки из Excel в DataFrame как пустые
        empty_columns = [col for col in existing_headers if col not in df_to_add.columns]
        if empty_columns:
            empty_df = pd.DataFrame({col: [""] * len(df_to_add) for col in empty_columns})
            df_to_add = pd.concat([df_to_add, empty_df], axis=1)

        # Упорядочим колонки строго как в Excel
        df_to_add = df_to_add[existing_headers]

        with pd.ExcelWriter(output_path, engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:
            # Таблица анализа
            df_to_add.to_excel(writer, sheet_name="Результаты анализа", index=False, startrow=startrow - 1, header=False)
        logger.info(f"Данные из файла '{doc_path.split('/')[-1]}' сохранены в Excel.")

